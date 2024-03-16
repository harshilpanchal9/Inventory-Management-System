from datetime import date, timedelta
import datetime
import decimal
from urllib import response
from django.db.models import Sum
import json
from django.shortcuts import get_object_or_404, render, redirect
from AdminPanel.models import AdminLogin, CategoryModel, ProductModel, Suppliers, Purchases, PurchasedProducts, Customers, Sells, SoldProducts, SupplierExpenses, SupplierExpensesProducts, CustomerExpenses, CustomerExpensesProducts, OtherExpensesOfPurchase, OtherExpensesOfSells, SuppliersDuePayments, SuppliersDueBills, CustomersDuePayments, CustomersDueBills
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment
import subprocess
from django.db.models import Q

# Create your views here.
def index(request):
    try:
        if request.method == "POST":
            email = request.POST.get('email')
            password = request.POST.get('password')        
            isValiduser = AdminLogin.ValidUser(request,email,password)  
            if isValiduser:  
                return redirect("/dashboard")
            else:
                return redirect("/")
    except isValiduser.DoesNotExist:
        isValiduser = None
    return render(request, "index.html")

def dashboard(request):
    if 'email' in request.session:
        Total_category = CategoryModel.objects.count()
        Total_product = ProductModel.objects.count()
        Total_purchases = Purchases.objects.count()
        Total_sells = Sells.objects.count()
        
        # To show 7 recent purchases
        Recent_purchases = Purchases.objects.all().order_by('-billNo')[:7]
        Recent_sales = Sells.objects.all().order_by('-billNo')[:7]
        # To show top selling product
        decimal.getcontext().prec = 2
        top_selling_products = ProductModel.objects.order_by('-quantity_sold')[:6]
        total_quantity_sold = sum(product.quantity_sold for product in top_selling_products)
        ratios = {}
        for product in top_selling_products:
            ratio = (product.quantity_sold / total_quantity_sold) * 100
            ratios[product.productname] = ratio

        # Calculate revenue for the last month in percentage

        # To find last month each day revenue
        sells = Sells.objects.all().values('billDate', 'grossTotal')
        sell_data = list(sells)

        # Get the current date and last month's date
        current_date = datetime.date.today()
        last_month_date = current_date.replace(day=1) - datetime.timedelta(days=1)
        last_to_last_month_date = last_month_date.replace(day=1) - datetime.timedelta(days=1)
        # print("lmd: ",last_month_date)
        # print("ltlmd: ",last_to_last_month_date)
        # Filter the sell_data to get the gross totals for the last month and the last to last month
        last_month_gross_totals = [float(sell['grossTotal']) for sell in sell_data if sell['billDate'].month == last_month_date.month]
        last_to_last_month_gross_totals = [sell['grossTotal'] for sell in sell_data if sell['billDate'].month == last_to_last_month_date.month]
        # print("lm gross: ",last_month_gross_totals)
        
        # print("ltlm gross: ",last_to_last_month_gross_totals)
        # Calculate the total revenue for the last month and the last to last month
        last_month_total_revenue = float(sum(last_month_gross_totals))
        last_to_last_month_total_revenue = sum(last_to_last_month_gross_totals)
        # print(last_month_total_revenue)
        # print(last_to_last_month_total_revenue)
        # Set the base value as the total revenue from the last to last month
        base_value = last_to_last_month_total_revenue
        # Calculate the percentage
        percentage = (last_month_total_revenue / base_value) * 100
        # Optionally, round the percentage to two decimal places
        revenue_percentage = round(percentage, 2)
        # Print or use the percentage value as needed
        # print(revenue_percentage)

        # To parse the sell data
        sell_data_chart = Sells.objects.all().values('billDate', 'grossTotal')
        # print("sell_data_chart",sell_data_chart)
        # Convert sell data to a list of dictionaries
        sell_data_list = list(sell_data_chart)
        # Format the date values as strings in each dictionary
        for sell in sell_data_list:
            sell['billDate'] = sell['billDate'].strftime('%Y-%m-%d')
        # Convert the list of dictionaries to a JSON string
        sell_data_json = json.dumps(sell_data_list)

        # Calculation for last month Net Profit
        purchases = Purchases.objects.all().values('billDate', 'grossTotal')
        purchase_data = list(purchases)
        # print("Purchase_data", purchase_data)
        last_month_gross_costs_total = [float(purchase['grossTotal']) for purchase in purchase_data if purchase['billDate'].month == last_month_date.month] 
        last_month_total_costs = float(sum(last_month_gross_costs_total))
        net_profit = float(last_month_total_revenue - last_month_total_costs)
        net_profit_percentage = ((last_month_total_revenue - last_month_total_costs) / last_month_total_revenue) * 100 if last_month_total_revenue else 0
        net_profit_percentage = round(net_profit_percentage, 2)

        # To parse the purchase data
        purchase_data_chart = Purchases.objects.all().values('billDate', 'grossTotal')
        # Convert sell data to a list of dictionaries
        purchase_data_list = list(purchase_data_chart)
        # Format the date values as strings in each dictionary
        for purchase in purchase_data_list:
            purchase['billDate'] = purchase['billDate'].strftime('%Y-%m-%d')
        # Convert the list of dictionaries to a JSON string
        purchase_data_json = json.dumps(purchase_data_list)

        context = { 'Total_category': Total_category,
                    'Total_product':Total_product, 
                    'Total_purchases': Total_purchases, 
                    'Total_sells':Total_sells, 
                    'Recent_purchases': Recent_purchases, 
                    'Recent_sales': Recent_sales, 
                    'ratios':ratios, 
                    #'revenue_percentage':revenue_percentage,
                    'last_month_total_revenue': last_month_total_revenue, # Total revenue in chart
                    'revenue_percentage':last_month_total_revenue, 
                    'revenue_sell_data':sell_data_json,             # For printing data in chart
                    #'net_profit_percentage':net_profit,
                    'last_month_net_profit':net_profit,                        # Total net_profit in chart
                    'net_profit_purchase_data':purchase_data_json   # For printing data in chart
                }
        return render(request, "dashboard.html",context)
    else:
        return redirect("/")

#------------- Start Category Section --------------
def addcategory(request):
    if 'email' in request.session:
        try:
            if request.method == "POST":
                CategoryModel.addCategory(request)
        except Exception as e:
            print(e)
        return render(request,"add-category.html")
    else:
        return redirect("/")

#------------- Delete Session --------------
def delSession(request):
    print(dict(request.session))
    request.session.flush()
    request.session.clear_expired()
    print(dict(request.session))
    return render(request,"index.html")

@csrf_exempt
def viewcategory(request):
    if 'email' in request.session:
        counter = CategoryModel.objects.count()

        # Pageload content & search functionality
        if request.method == 'GET':
            pageload = request.GET.get('pageload')
            if pageload:
                categories = CategoryModel.objects.all().order_by('categoryname')
                html = render_to_string('view-category.html', {'categories': categories})
                data = {'html': html, 'categories': list(categories.values()),'category_counter':counter}
                return JsonResponse(data)
            category_name = request.GET.get('search-input')
            if category_name:
                categories = CategoryModel.objects.filter(categoryname__icontains=category_name).order_by('categoryname')
                html = render_to_string('view-category.html', {'categories': categories})
                data = {'html': html, 'categories': list(categories.values()),'category_counter':counter}
                return JsonResponse(data)

        if request.method == "POST":
            categoryname = request.POST.get('categoryname')
            # print("categoryname:",categoryname)
            result = CategoryModel.deleteCategory(request)
            if result:
                flag = 0
                return render(request,"view-category.html",{'function_call':"call",'delete_result':flag,'categoryname':categoryname,'category_counter':counter})
            else:
                flag = 1
                return render(request,"view-category.html",{'delete_result':flag,'categoryname':categoryname,'category_counter':counter})

        return render(request,"view-category.html")
    else:
        return redirect("/")

def editcategory(request):
    if 'email' in request.session:
        try:
            if request.method == "POST":
                result = CategoryModel.updateCategory(request)
                if result:
                    messages.success(request, "Category updated successfully.")
                    return redirect('/view-category')
                else:
                    messages.warning(request, "Cannot update Category, Category name already exists.")
                    return redirect('/view-category')
        except Exception as e:
            print(e)
        
        categoryname = request.GET.get('categoryname')
        context = {'categoryname': categoryname}
        return render(request,"edit-category.html",context)
    else:
        return redirect("/")

#------------- End Category Section ---------------
 
#------------- Start Product Section --------------
def addproduct(request):
    if 'email' in request.session:
        try:
            if request.method == "POST":
                ProductModel.addProduct(request)
        except Exception as e:
            print(e)

        # For showing all available products in category Dropdown

        all_categories = CategoryModel.objects.all()
        all_products = ProductModel.objects.all()
        if all_products:
            lastProductId = ProductModel.objects.all().last().__getattribute__("productid")
            return render(request,"add-product.html",{'all_categories':all_categories,'newProductId': lastProductId,'flag':0})
        else:
            return render(request,"add-product.html",{'all_categories':all_categories,'flag':2})
    else:
        return redirect("/")

@csrf_exempt
def viewproducts(request):
    if 'email' in request.session:
        if request.method == 'GET':
            Category_Data = CategoryModel.objects.all()
            pageload = request.GET.get('pageload')
            if pageload:
                products = ProductModel.objects.all().order_by('-productid')
                html = render_to_string('view-products.html', {'products': products})
                data = {'html': html, 'products': list(products.values()), 'Category_Data': list(Category_Data.values())}
                return JsonResponse(data)
            product_query = request.GET.get('search-input')
            if product_query:
                products = ProductModel.objects.filter(Q(productname__icontains=product_query) | Q(productid__icontains=product_query)).order_by('-productid')
                html = render_to_string('view-products.html', {'products': products})
                data = {'html': html, 'products': list(products.values()), 'Category_Data': list(Category_Data.values())}
                return JsonResponse(data)

        if request.method == "POST":
            productname = request.POST.get('productname')
            result = ProductModel.deleteProduct(request)
            if result:
                flag = 0
                return render(request,"view-products.html",{'function_call':"call",'delete_result':flag,'productname':productname})
            else:
                flag = 1
                return render(request,"view-products.html",{'delete_result':flag,'productname':productname})
        return render(request,"view-products.html")
    else:
        return redirect("/")

def editproduct(request):
    if 'email' in request.session:
        try:
            flag = ProductModel.updateProduct(request)
            if flag == 0:
                messages.success(request, "Product details updated successfully.")
                return redirect('/view-products')
            
        except Exception as e:
            print(e)
        
        productid = request.GET.get('productid')
        category = request.GET.get('category')
        productname = request.GET.get('productname')
        productsize = request.GET.get('productsize')
        productcolor = request.GET.get('productcolor')
        productmaterial = request.GET.get('productmaterial')
        productstock = request.GET.get('productstock')
        purchaseprice = request.GET.get('purchaseprice')
        sellprice = request.GET.get('sellprice')
        productdesc = request.GET.get('productdesc')
        productimage = request.GET.get('productimage')
        Category_Data = CategoryModel.objects.values()
        context = {'Category_Data':Category_Data,'productid':productid,'category':category,'productname': productname,'productsize':productsize,'productcolor':productcolor,'productmaterial':productmaterial,'productstock':productstock,'purchaseprice':purchaseprice, 'sellprice':sellprice,'productdesc':productdesc,'productimage':productimage}
        return render(request,"edit-product.html",context)
    else:
        return redirect("/")

#------------- End Product Section --------------

#------------- Purchase Section -----------------

def addPurchase(request):
    if 'email' in request.session:
        if request.method == 'GET':
            currentDate = date.today()
            suppliers = Suppliers.objects.all()
            billNo = Purchases.generateBillNo()

            # To open a calculator using ajax
            calculator = request.GET.get('calculator')
            if calculator:
                subprocess.Popen('calc.exe')

            context = {'currentDate': currentDate, 'suppliers': suppliers, 'billNo' : billNo}
            return render(request,"add-purchase.html", context)
        
        elif request.method == 'POST':
            # Generating purchase bill data
            purchasebill = Purchases()
            isPurchaseBillGenerated = purchasebill.addPurchase(request)

            # Adding products
            if isPurchaseBillGenerated:
                for key, value in request.POST.items():
                    if key.startswith('product'):
                        rowNumber = key.split('_')[-1]
                        purchasedProduct = PurchasedProducts()
                        isPurchasedProductAdded = purchasedProduct.addPurchasedProduct(request, rowNumber)

                for key, value in request.POST.items():
                    if key.startswith('expenseType'):
                        expenseRowNumber = key.split('_')[-1]
                        OtherExpense = OtherExpensesOfPurchase()

                        if request.POST.get(f'expenseAmount_{expenseRowNumber}') != '':
                            OtherExpense.addExpense(request, expenseRowNumber)
            
            if isPurchaseBillGenerated and isPurchasedProductAdded:
                messages.success(request, 'Bill registered successfully')
                return redirect('/add-purchase')
            else:
                messages.warning(request, 'Bill not registered.')
                return redirect('/add-purchase')
    else:
        return redirect("/")

def purchases(request):
    if 'email' in request.session:
        if request.method == 'GET':
            purchases = Purchases.objects.all().order_by('-billNo')
            context = {'purchases': purchases}
            return render(request,"purchases.html", context)
        if request.method == 'POST':
            billNo = request.POST.get('billNo')
            
            # Retrieve the existing payment status from the database
            purchase = Purchases.objects.get(billNo=billNo)
            current_status = purchase.paymentStatus
            
            # Toggle the payment status
            if current_status == "Partial Credit":
                new_status = "Paid"
            elif current_status == "Paid":
                new_status = "Paid"
            elif current_status == "Full Credit":
                new_status = "Paid"
            elif current_status == "Return":
                new_status = "Return"
            # Update the payment status in the database
            Purchases.objects.filter(billNo=billNo).update(paymentStatus=new_status)
            purchases = Purchases.objects.all().order_by('-billNo') 
            context = {'purchases': purchases}
            return render(request,"purchases.html", context)
    else:
        return redirect("/")

def viewPurchaseBill(request):
    if 'email' in request.session:
        if request.method == 'GET':
            purchase = Purchases.objects.get(billNo=request.GET.get('billNo', None))
            purchasedProducts = list()
            purchasedProducts = PurchasedProducts.objects.filter(billNo=request.GET.get('billNo', None))
            otherExpenses = OtherExpensesOfPurchase.objects.filter(billNo=request.GET.get('billNo', None))
            context = {'purchase': purchase, 'purchasedProducts':purchasedProducts, 'otherExpenses': otherExpenses}
            return render(request, 'view-purchase-bill.html', context) 
    else:
        return redirect("/")
    
@csrf_exempt
def purchaseReturn(request):
    if 'email' in request.session:
        purchases = Purchases.objects.all().order_by('-billNo') 
        context = {'purchases': purchases}
        if request.method == 'POST':
            billNo = request.POST.get('billNo')
            purchase = Purchases.objects.get(billNo=billNo)
            current_status = purchase.paymentStatus
            if current_status == "Partial Credit":
                new_status = "Return"
            elif current_status == "Paid":
                new_status = "Return"
            elif current_status == "Full Credit":
                new_status = "Return"
            elif current_status == "Return":
                new_status = "Return"
            Purchases.objects.filter(billNo=billNo).update(paymentStatus=new_status)
            purchases = Purchases.objects.all().order_by('-billNo') 
            html = render_to_string('purchase-return.html', {'purchases': purchases})
            data = {'html': html}
            return JsonResponse(data)
        return render(request,"purchase-return.html",context)
    else:
        return redirect("/")

#------------- Supplier Section -----------------

def suppliers(request):
    if 'email' in request.session:
        suppliers = Suppliers.objects.all().order_by('-supplierId')
        context = {'suppliers': suppliers}
        return render(request,"suppliers.html", context) 
    else:
        return redirect("/")
                                                                                                                                                                                                                                                                                                                      
def addSupplier(request):
    if 'email' in request.session:
        if request.method == 'POST':
            sId = request.POST.get('supplierId')
            sName = request.POST.get('supplierName')
            sGst = request.POST.get('supplierGst')
            sAddress = request.POST.get('supplierAddress')
            sEmail = request.POST.get('supplierEmail')
            sPhone = request.POST.get('supplierPhone')

            supplier = Suppliers()
            isSupplierAdded = supplier.addSupplier(supplierId=sId, supplierName=sName, supplierGst=sGst, supplierAddress=sAddress, supplierEmail=sEmail, supplierPhone=sPhone)

            if isSupplierAdded:
                messages.success(request, "Supplier added successfully.")
            else:
                messages.warning(request, "Supplier not added.")
            return redirect('/add-supplier')
        
        elif request.method == 'GET':
            supplierId = Suppliers.generateSupplierId()
            context = {'supplierId' : supplierId}
            return render(request,"add-supplier.html", context)
    else:
        return redirect("/")

def editSupplier(request):
    if 'email' in request.session:
        if request.method == "POST":
            supplierId = request.POST.get("supplierId")
            print("post sid: ",supplierId)
            supplierName = request.POST.get("supplierName")
            supplierGst = request.POST.get("supplierGst")
            supplierAddress = request.POST.get("supplierAddress")
            supplierEmail = request.POST.get("supplierEmail")
            supplierPhone = request.POST.get("supplierPhone")
            Suppliers.objects.filter(pk=supplierId).update(
                supplierName = supplierName,
                supplierGst = supplierGst,
                supplierAddress = supplierAddress,
                supplierEmail = supplierEmail,
                supplierPhone = supplierPhone,
            )
            messages.success(request, "Supplier details updated successfully.")
            return redirect("/suppliers")

        supplierId = request.GET.get('supplierId')
        supplierName = request.GET.get("supplierName")
        supplierGst = request.GET.get("supplierGst")
        supplierAddress = request.GET.get("supplierAddress")
        supplierEmail = request.GET.get("supplierEmail")
        supplierPhone = request.GET.get("supplierPhone")
        context = {'supplierId':supplierId,'supplierName':supplierName,'supplierGst':supplierGst,'supplierAddress':supplierAddress,'supplierEmail':supplierEmail,'supplierPhone':supplierPhone}
        return render(request, 'edit-supplier.html',context)
    else:
        return redirect("/")

def getSupplierData(request):
    if 'email' in request.session:
        supplierId = request.GET.get('supplierId', None)
        
        data = dict()
        if supplierId:
            supplier = Suppliers.objects.get(supplierId=supplierId)
            data['phone'] = supplier.supplierPhone
            data['gst'] = supplier.supplierGst
        return JsonResponse(data)
    else:
        return redirect("/")
    
def supplierExpenses(request):
    if 'email' in request.session:
        if request.method == 'GET':
            currentDate = date.today()
            suppliers = Suppliers.objects.all()
            billNo = SupplierExpenses.generateBillNo()

            # To open a calculator using ajax
            calculator = request.GET.get('calculator')
            if calculator:
                subprocess.Popen('calc.exe')

            context = {'currentDate': currentDate, 'suppliers': suppliers, 'billNo' : billNo}
            return render(request, 'supplier-expenses.html', context)
        
        elif request.method == 'POST':
            # Generating purchase bill data
            supplierExpensesbill = SupplierExpenses()
            isSupplierExpensesBillGenerated = supplierExpensesbill.addSupplierExpenses(request)

            # Adding products
            if isSupplierExpensesBillGenerated:
                for key, value in request.POST.items():
                    if key.startswith('product'):
                        rowNumber = key.split('_')[-1]
                        supplierExpensesProducts = SupplierExpensesProducts()
                        isSupplierExpensesProductAdded = supplierExpensesProducts.addSupplierExpensesProduct(request, rowNumber)
            
            if isSupplierExpensesBillGenerated and isSupplierExpensesProductAdded:
                messages.success(request, 'Expense bill registered successfully')
                return redirect('/supplier-expenses')
            else:
                messages.warning(request, 'Expense bill not registered.')
                return redirect('/supplier-expenses')
    else:
        return redirect("/")

def viewSupplierExpenses(request):
    if 'email' in request.session:
        if request.method == 'GET':
            supplierExpenses = SupplierExpenses.objects.all().order_by('-billNo')
            context = {'supplierExpenses': supplierExpenses}
            return render(request, "view-supplier-expenses.html", context)
    else:
        return redirect("/")
    
def viewSupplierExpenseBill(request):
    if 'email' in request.session:
        if request.method == 'GET':
            supplierExpenses = SupplierExpenses.objects.get(billNo=request.GET.get('billNo', None))
            supplierExpensesProducts = list()
            supplierExpensesProducts = SupplierExpensesProducts.objects.filter(billNo=request.GET.get('billNo', None))
            context = {'supplierExpenses': supplierExpenses, 'supplierExpensesProducts':supplierExpensesProducts}
            return render(request, 'view-supplier-expense-bill.html', context) 
    else:
        return redirect("/")

#------------- Sell Section -----------------
@csrf_exempt
def addSell(request):
    if 'email' in request.session:
        if request.method == 'GET':
            currentDate = date.today()
            customers = Customers.objects.all()
            all_products = ProductModel.objects.all()
            products_json = list(ProductModel.objects.values_list('productname', flat=True))
            products = json.dumps(products_json)
            billNo = Sells.generateBillNo()
            sell = request.GET.get('sell')

            # To open a calculator using ajax
            calculator = request.GET.get('calculator')
            if calculator:
                subprocess.Popen('calc.exe')

            # Ajax code for getting product stock
            productStock = request.GET.get("productStock")
            productName = request.GET.get("productName")
            if productStock == "productStock":
                product = ProductModel.objects.get(productname=productName)
                stock = product.productstock
                data = {'stock': stock}
                return JsonResponse(data)
            context = {'currentDate': currentDate, 'customers': customers, 'products':products,'all_products':all_products, 'billNo' : billNo, 'sell': sell}
            return render(request,"add-sell.html", context)
        
        elif request.method == 'POST':
            # Generating sell bill data
            sellbill = Sells()
            isSellBillGenerated = sellbill.addSell(request)

            # Adding products
            if isSellBillGenerated:
                for key, value in request.POST.items():
                    if key.startswith('product'):
                        rowNumber = key.split('_')[-1]
                        soldProduct = SoldProducts()
                        isSoldProductAdded = soldProduct.addSoldProduct(request, rowNumber)

                for key, value in request.POST.items():
                    if key.startswith('expenseType'):
                        expenseRowNumber = key.split('_')[-1]
                        OtherExpense = OtherExpensesOfSells()
                        if request.POST.get(f'expenseAmount_{expenseRowNumber}') != '':
                            OtherExpense.addExpense(request, expenseRowNumber)
            
            if isSellBillGenerated and isSoldProductAdded:
                messages.success(request, 'Bill registered successfully')
                return redirect('/add-sell')
            else:
                messages.warning(request, 'Bill not registered.')
                return redirect('/add-sell')
    else:
        return redirect("/")

def sells(request):
    if 'email' in request.session:
        if request.method == 'GET':
            sells = Sells.objects.all().order_by('-billNo')
            context = {'sells': sells}
            return render(request,"sells.html", context)
        if request.method == 'POST':
            billNo = request.POST.get('billNo')
            
            # Retrieve the existing payment status from the database
            sell = Sells.objects.get(billNo=billNo)
            current_status = sell.paymentStatus
            
            # Toggle the payment status
            if current_status == "Partial Credit":
                new_status = "Paid"
            elif current_status == "Paid":
                new_status = "Paid"
            elif current_status == "Full Credit":
                new_status = "Paid"
            elif current_status == "Return":
                new_status = "Return"
            # Update the payment status in the database
            Sells.objects.filter(billNo=billNo).update(paymentStatus=new_status)
            sells = Sells.objects.all().order_by('-billNo') 
            context = {'sells': sells}
            return render(request,"sells.html", context)
    else:
        return redirect("/")
             
def viewSellBill(request):
    if 'email' in request.session:
        if request.method == 'GET':
            sell = Sells.objects.get(billNo=request.GET.get('billNo', None))
            soldProducts = list()
            soldProducts = SoldProducts.objects.filter(billNo=request.GET.get('billNo', None))
            otherExpenses = OtherExpensesOfSells.objects.filter(billNo=request.GET.get('billNo', None))
            print(soldProducts)
            context = {'sell': sell, 'soldProducts':soldProducts, 'otherExpenses': otherExpenses}
            return render(request, 'view-sell-bill.html', context) 
    else:
        return redirect("/")
    
@csrf_exempt
def sellReturn(request):
    if 'email' in request.session:
        sells = Sells.objects.all().order_by('-billNo') 
        context = {'sells': sells}
        if request.method == 'POST':
            billNo = request.POST.get('billNo')
            sell = Sells.objects.get(billNo=billNo)
            current_status = sell.paymentStatus
            if current_status == "Partial Credit":
                new_status = "Return"
            elif current_status == "Paid":
                new_status = "Return"
            elif current_status == "Full Credit":
                new_status = "Return"
            elif current_status == "Return":
                new_status = "Return"
            Sells.objects.filter(billNo=billNo).update(paymentStatus=new_status)
            sells = Sells.objects.all().order_by('-billNo') 
            html = render_to_string('sell-return.html', {'sells': sells})
            data = {'html': html}
            return JsonResponse(data)
        return render(request,"sell-return.html",context)
    else:
        return redirect("/")

#------------- Customer Section -----------------

def customers(request):
    if 'email' in request.session:
        customers = Customers.objects.all().order_by('-customerId')
        context = {'customers': customers}
        return render(request,"customers.html", context)  
    else:
        return redirect("/")
               
def addCustomer(request):
    if 'email' in request.session:
        if request.method == 'POST':
            sId = request.POST.get('customerId')
            sName = request.POST.get('customerName')
            sGst = request.POST.get('customerGst')
            sAddress = request.POST.get('customerAddress')
            sEmail = request.POST.get('customerEmail')
            sPhone = request.POST.get('customerPhone')

            customer = Customers()
            isCustomerAdded = customer.addCustomer(customerId=sId, customerName=sName, customerGst=sGst, customerAddress=sAddress, customerEmail=sEmail, customerPhone=sPhone)

            if isCustomerAdded:
                messages.success(request, "Customer added successfully.")
            else:
                messages.warning(request, "Customer not added.")
            return redirect('/add-customer')
        
        elif request.method == 'GET':
            customerId = Customers.generateCustomerId()
            context = {'customerId' : customerId}
            return render(request,"add-customer.html", context)
    else:
        return redirect("/")

def editCustomer(request):
    if 'email' in request.session:
        if request.method == "POST":
            customerId = request.POST.get("customerId")
            print("post sid: ",customerId)
            customerName = request.POST.get("customerName")
            customerGst = request.POST.get("customerGst")
            customerAddress = request.POST.get("customerAddress")
            customerEmail = request.POST.get("customerEmail")
            customerPhone = request.POST.get("customerPhone")
            Customers.objects.filter(pk=customerId).update(
                customerName = customerName,
                customerGst = customerGst,
                customerAddress = customerAddress,
                customerEmail = customerEmail,
                customerPhone = customerPhone,
            )
            messages.success(request, "Customer details updated successfully.")
            return redirect("/customers")

        customerId = request.GET.get('customerId')
        print("customerId",customerId)
        customerName = request.GET.get("customerName")
        customerGst = request.GET.get("customerGst")
        customerAddress = request.GET.get("customerAddress")
        customerEmail = request.GET.get("customerEmail")
        customerPhone = request.GET.get("customerPhone")
        context = {'customerId':customerId,'customerName':customerName,'customerGst':customerGst,'customerAddress':customerAddress,'customerEmail':customerEmail,'customerPhone':customerPhone}
        return render(request, 'edit-customer.html',context)
    else:
        return redirect("/")

def getCustomerData(request):
    if 'email' in request.session:
        customerId = request.GET.get('customerId', None)
        
        data = dict()
        if customerId:
            customer = Customers.objects.get(customerId=customerId)
            data['phone'] = customer.customerPhone
            data['gst'] = customer.customerGst
        return JsonResponse(data)
    else:
        return redirect("/")
    
def customerExpenses(request):
    if 'email' in request.session:
        if request.method == 'GET':
            currentDate = date.today()
            customers = Customers.objects.all()
            billNo = CustomerExpenses.generateBillNo()

            # To open a calculator using ajax
            calculator = request.GET.get('calculator')
            if calculator:
                subprocess.Popen('calc.exe')

            context = {'currentDate': currentDate, 'customers': customers, 'billNo' : billNo}
            return render(request, 'customer-expenses.html', context)
        
        elif request.method == 'POST':
            # Generating sell bill data
            customerExpensesbill = CustomerExpenses()
            isCustomerExpensesBillGenerated = customerExpensesbill.addCustomerExpenses(request)

            # Adding products
            if isCustomerExpensesBillGenerated:
                for key, value in request.POST.items():
                    if key.startswith('product'):
                        rowNumber = key.split('_')[-1]
                        customerExpensesProducts = CustomerExpensesProducts()
                        isCustomerExpensesProductAdded = customerExpensesProducts.addCustomerExpensesProduct(request, rowNumber)
                        print(f'{key}: {value}')
                        print(rowNumber)
            
            if isCustomerExpensesBillGenerated and isCustomerExpensesProductAdded:
                messages.success(request, 'Expense bill registered successfully')
                return redirect('/customer-expenses')
            else:
                messages.warning(request, 'Expense bill not registered.')
                return redirect('/customer-expenses')
    else:
        return redirect("/")

def viewCustomerExpenses(request):
    if 'email' in request.session:
        if request.method == 'GET':
            customerExpenses = CustomerExpenses.objects.all().order_by('-billNo')
            context = {'customerExpenses': customerExpenses}
        return render(request, "view-customer-expenses.html", context)
    else:
        return redirect("/")

def viewCustomerExpenseBill(request):
    if 'email' in request.session:
        if request.method == 'GET':
            customerExpenses = CustomerExpenses.objects.get(billNo=request.GET.get('billNo', None))
            customerExpensesProducts = list()
            customerExpensesProducts = CustomerExpensesProducts.objects.filter(billNo=request.GET.get('billNo', None))
            context = {'customerExpenses': customerExpenses, 'customerExpensesProducts':customerExpensesProducts}
        return render(request, "view-customer-expense-bill.html", context)
    else:
        return redirect("/")

#------------- Supplier Due payments -----------------
def suppliersDuePayments(request):
    allDuePayments = SuppliersDuePayments.getDuePayments()
    context = {'allDuePayments': allDuePayments}
    return render(request,"suppliers-due-payments.html", context)

#------------- Supplier Due Bills -----------------
def suppliersDueBills(request):
    if request.method == 'POST':
        billNo = request.POST.get('billNo')
        supplierId = request.POST.get('supplierId')
        isDuesCleared = SuppliersDueBills.clearDues(billNo, supplierId)

        if isDuesCleared:
            print("duesCleared")
            allDueBills, supplierName = SuppliersDueBills.getDueBills(supplierId)
            messages.success(request, 'Due Invoice Cleared')
            context = {'allDueBills': allDueBills, 'supplierName': supplierName}
        return render(request,"suppliers-due-bills.html", context)
    else:
        # Handle other HTTP methods or return an error response if needed
        supplierId = request.GET.get('supplierId')
        allDueBills, supplierName = SuppliersDueBills.getDueBills(supplierId)
        # if all
        context = {'allDueBills': allDueBills, 'supplierName': supplierName}
        return render(request,"suppliers-due-bills.html", context)

#------------- Customer Due payments -----------------
def customersDuePayments(request):
    allDuePayments = CustomersDuePayments.getDuePayments()
    context = {'allDuePayments': allDuePayments}
    return render(request,"customers-due-payments.html", context)

#------------- Customer Due Bills -----------------.
def customersDueBills(request):
    if request.method == 'POST':
        billNo = request.POST.get('billNo')
        customerId = request.POST.get('customerId')
        isDuesCleared = CustomersDueBills.clearDues(billNo, customerId)

        if isDuesCleared:
            print("duesCleared")
            allDueBills, customerName = CustomersDueBills.getDueBills(customerId)
            messages.success(request, 'Due Invoice Cleared')
            context = {'allDueBills': allDueBills, 'customerName': customerName}
        return render(request,"customers-due-bills.html", context)
    else:
        customerId = request.GET.get('customerId')
        allDueBills, customerName = CustomersDueBills.getDueBills(customerId)
        context = {'allDueBills': allDueBills, 'customerName': customerName}
        return render(request,"customers-due-bills.html", context)

#------------- Invoice Section -----------------

def Sellinvoice(request):
    if 'email' in request.session:
        if request.method == "GET":
            sellBillNo = request.GET.get("sell-billNo")
            print("sellbill: ",sellBillNo)
            sellData = get_object_or_404(Sells, pk=sellBillNo)
            customerDetails = get_object_or_404(Customers, customerName=sellData.customerName)
            soldProducts = SoldProducts.objects.filter(billNo=sellBillNo)
            otherExpensesData = OtherExpensesOfSells.objects.filter(billNo=sellBillNo)
            context = {'sellData':sellData, 'customerDetails':customerDetails, 'soldProducts': soldProducts, 'otherExpensesData': otherExpensesData}
            return render(request,"sell-invoice.html",context)
    else:
        return redirect("/")

def Purchaseinvoice(request):
    if 'email' in request.session:
        if request.method == "GET":
            purchaseBillNo = request.GET.get("purchase-billNo")
            purchaseData = get_object_or_404(Purchases, pk=purchaseBillNo)
            supplierDetails = get_object_or_404(Suppliers, supplierName=purchaseData.supplierName)
            purchasedProducts = PurchasedProducts.objects.filter(billNo=purchaseBillNo)
            otherExpensesData = OtherExpensesOfPurchase.objects.filter(billNo=purchaseBillNo)
            context = {'purchaseData':purchaseData, 'supplierDetails':supplierDetails, 'purchasedProducts': purchasedProducts, 'otherExpensesData': otherExpensesData}
            return render(request,"purchase-invoice.html",context)
    else:
        return redirect("/")
    
# To generate excel file of sell invoice
def generate_excel_for_sell(request):
    if 'email' in request.session:
        sellBillNo = request.GET.get("sell-billNo")
        soldProducts = SoldProducts.objects.filter(billNo=sellBillNo)
        sellBillData = Sells.objects.filter(billNo=sellBillNo)
        otherExpensesData = OtherExpensesOfSells.objects.filter(billNo=sellBillNo)

        wb = Workbook()

        sheet = wb.active

        sheet['A1'] = 'BillNo.'
        sheet['B1'] = 'Description'
        sheet['C1'] = 'Quantity'
        sheet['D1'] = 'Unit'
        sheet['E1'] = 'Price/unit'
        sheet['F1'] = 'Amount'

        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center')

        row_index = 2

        # To print the product data of invoice 
        for index, soldProduct in enumerate(soldProducts, start=1):
            sheet.cell(row=row_index, column=1, value=sellBillNo)
            sheet.cell(row=row_index, column=2, value=soldProduct.product)
            sheet.cell(row=row_index, column=3, value=soldProduct.quantity)
            sheet.cell(row=row_index, column=4, value=soldProduct.unit)
            sheet.cell(row=row_index, column=5, value=soldProduct.pricePerUnit)
            sheet.cell(row=row_index, column=6, value=soldProduct.amount)
            row_index += 1

        # To print the Other Expenses data of invoice
        align = Alignment(horizontal='center')
        left = Alignment(horizontal='left')
        row_index += 1
        sheet.cell(row=row_index, column=1, value='No.').alignment = left
        sheet.cell(row=row_index, column=2, value='Expense type').alignment = align
        sheet.cell(row=row_index, column=3, value='Amount').alignment = align
        row_index += 1

        for index, otherExpensesList in enumerate(otherExpensesData, start=1):
            sheet.cell(row=row_index, column=1, value=index).alignment = left
            sheet.cell(row=row_index, column=2, value=otherExpensesList.expenseType)
            sheet.cell(row=row_index, column=3, value=otherExpensesList.amount)
            row_index += 1

        row_index += 1
        for index, sellBill in enumerate(sellBillData, start=1):
            sheet.cell(row=row_index, column=4, value='Total')
            sheet.cell(row=row_index, column=6, value=sellBill.total)
            row_index += 1

            if sellBill.gst:
                sheet.cell(row=row_index, column=4, value='GST 18%')
                sheet.cell(row=row_index, column=6, value=sellBill.gst)
                row_index += 1

            if sellBill.otherExpenses != 0.0:
                sheet.cell(row=row_index, column=4, value='Total Other Expenses')
                sheet.cell(row=row_index, column=6, value=sellBill.otherExpenses)
                row_index += 1

            sheet.cell(row=row_index, column=4, value='Gross Total')
            sheet.cell(row=row_index, column=6, value=sellBill.grossTotal)
            row_index += 1

            if sellBill.discount != 0.0:
                sheet.cell(row=row_index, column=4, value='Discount')
                sheet.cell(row=row_index, column=6, value=sellBill.discount)
                row_index += 1

            if sellBill.paidAmount != 0.0:
                sheet.cell(row=row_index, column=4, value='Paid amount')
                sheet.cell(row=row_index, column=6, value=sellBill.paidAmount)
                row_index += 1

            sheet.cell(row=row_index, column=4, value='Net total')
            sheet.cell(row=row_index, column=6, value=sellBill.netTotal)

        sheet.column_dimensions['B'].width = 30

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Sell_invoice.xlsx"'

        wb.save(response)

        return response
    else:
        return redirect("/")

# To generate excel file of Purchase invoice
def generate_excel_for_purchase(request):
    if 'email' in request.session:
        purchaseBillNo = request.GET.get("purchase-billNo")
        purchasedProducts = PurchasedProducts.objects.filter(billNo=purchaseBillNo)
        purchaseBillData = Purchases.objects.filter(billNo=purchaseBillNo)
        otherExpensesData = OtherExpensesOfPurchase.objects.filter(billNo=purchaseBillNo)

        wb = Workbook()

        sheet = wb.active

        sheet['A1'] = 'BillNo.'
        sheet['B1'] = 'Description'
        sheet['C1'] = 'Quantity'
        sheet['D1'] = 'Unit'
        sheet['E1'] = 'Price/unit'
        sheet['F1'] = 'Amount'

        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center')

        row_index = 2

        # To print the product data of invoice 
        for index, purchasedProduct in enumerate(purchasedProducts, start=1):
            sheet.cell(row=row_index, column=1, value=purchaseBillNo)
            sheet.cell(row=row_index, column=2, value=purchasedProduct.product)
            sheet.cell(row=row_index, column=3, value=purchasedProduct.quantity)
            sheet.cell(row=row_index, column=4, value=purchasedProduct.unit)
            sheet.cell(row=row_index, column=5, value=purchasedProduct.pricePerUnit)
            sheet.cell(row=row_index, column=6, value=purchasedProduct.amount)
            row_index += 1

        # To print the Other Expenses data of invoice
        align = Alignment(horizontal='center')
        left = Alignment(horizontal='left')
        row_index += 1
        sheet.cell(row=row_index, column=1, value='No.').alignment = left
        sheet.cell(row=row_index, column=2, value='Expense type').alignment = align
        sheet.cell(row=row_index, column=3, value='Amount').alignment = align
        row_index += 1

        for index, otherExpensesList in enumerate(otherExpensesData, start=1):
            sheet.cell(row=row_index, column=1, value=index).alignment = left
            sheet.cell(row=row_index, column=2, value=otherExpensesList.expenseType)
            sheet.cell(row=row_index, column=3, value=otherExpensesList.amount)
            row_index += 1

        row_index += 1

        for index, purchaseBill in enumerate(purchaseBillData, start=1):
            sheet.cell(row=row_index, column=4, value='Total')
            sheet.cell(row=row_index, column=6, value=purchaseBill.total)
            row_index += 1

            if purchaseBill.gst:
                sheet.cell(row=row_index, column=4, value='GST 18%')
                sheet.cell(row=row_index, column=6, value=purchaseBill.gst)
                row_index += 1

            if purchaseBill.otherExpenses != 0.0:
                sheet.cell(row=row_index, column=4, value='Total Other Expenses')
                sheet.cell(row=row_index, column=6, value=purchaseBill.otherExpenses)
                row_index += 1

            sheet.cell(row=row_index, column=4, value='Gross Total')
            sheet.cell(row=row_index, column=6, value=purchaseBill.grossTotal)
            row_index += 1

            if purchaseBill.discount != 0.0:
                sheet.cell(row=row_index, column=4, value='Discount')
                sheet.cell(row=row_index, column=6, value=purchaseBill.discount)
                row_index += 1

            if purchaseBill.paidAmount != 0.0:
                sheet.cell(row=row_index, column=4, value='Paid amount')
                sheet.cell(row=row_index, column=6, value=purchaseBill.paidAmount)
                row_index += 1

            sheet.cell(row=row_index, column=4, value='Net total')
            sheet.cell(row=row_index, column=6, value=purchaseBill.netTotal)

        sheet.column_dimensions['B'].width = 30

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Purchase_invoice.xlsx"'
        wb.save(response)
        return response
    else:
        return redirect("/")

# To generate excel file of Sell invoice data
def generate_excel_for_sellBills(request):
    if 'email' in request.session:
        SellData = Sells.objects.all()

        wb = Workbook()

        sheet = wb.active

        sheet['A1'] = 'BillNo.'                         #1
        sheet['B1'] = 'Bill Date.'                      #2
        sheet['C1'] = 'Customer Name'                   #3
        sheet['D1'] = 'Customer Phone'                  #4
        sheet['E1'] = 'Customer GST'                    #5
        sheet['F1'] = 'Product Name'                    #6
        sheet['G1'] = 'Quantity'                        #7
        sheet['H1'] = 'Unit'                            #8
        sheet['I1'] = 'Price/Unit'                      #9
        sheet['J1'] = 'Payment Type'                    #10
        sheet['K1'] = 'Payment Status'                  #11
        sheet['L1'] = 'Expense Type'                    #12
        sheet['M1'] = 'Expense Amount'                  #13
        sheet['N1'] = 'Total'                           #14
        sheet['O1'] = 'GST (18%)'                       #15
        sheet['P1'] = 'Total Of Other Expenses'         #16
        sheet['Q1'] = 'Gross Total'                     #17
        sheet['R1'] = 'Discount'                        #18
        sheet['S1'] = 'Paid Amount'                     #19
        sheet['T1'] = 'Net Total'                       #20
        sheet['U1'] = 'Description'                     #21

        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center')

        row_index = 2
        left = Alignment(horizontal='left')
        for index, sell in enumerate(SellData, start=1):
            formatted_date = sell.billDate.strftime('%Y-%m-%d')
            sheet.cell(row=row_index, column=1, value=sell.billNo).alignment = left
            sheet.cell(row=row_index, column=2, value=formatted_date)
            sheet.cell(row=row_index, column=3, value=sell.customerName)
            sheet.cell(row=row_index, column=4, value=sell.customerPhone)
            sheet.cell(row=row_index, column=5, value=sell.customerGst)

            product_column = 6

            sold_products = SoldProducts.objects.filter(billNo=sell.billNo)
            sold_products_length = len(sold_products)

            for i, product in enumerate(sold_products, start=1):
                sheet.cell(row=row_index, column=product_column, value=product.product)
                sheet.cell(row=row_index, column=product_column + 1, value=product.quantity)
                sheet.cell(row=row_index, column=product_column + 2, value=product.unit)
                sheet.cell(row=row_index, column=product_column + 3, value=product.pricePerUnit)
                if i == sold_products_length:
                    pass
                else:
                    row_index += 1
            sheet.cell(row=row_index, column=10, value=sell.paymentType)
            sheet.cell(row=row_index, column=11, value=sell.paymentStatus)

            other_expense = 12
            otherExpenses = OtherExpensesOfSells.objects.filter(billNo=sell.billNo)
            otherExpenses_length = len(otherExpenses)
            for i, otherExpense in enumerate(otherExpenses, start=1):
                sheet.cell(row=row_index, column=other_expense, value=otherExpense.expenseType)
                sheet.cell(row=row_index, column=other_expense + 1, value=otherExpense.amount)
                if i == otherExpenses_length:
                    pass
                else:
                    row_index += 1
            sheet.cell(row=row_index, column=14, value=sell.total)
            sheet.cell(row=row_index, column=15, value=sell.gst)
            sheet.cell(row=row_index, column=16, value=sell.otherExpenses)
            sheet.cell(row=row_index, column=17, value=sell.grossTotal)
            sheet.cell(row=row_index, column=18, value=sell.discount)
            sheet.cell(row=row_index, column=19, value=sell.paidAmount)
            sheet.cell(row=row_index, column=20, value=sell.netTotal)
            sheet.cell(row=row_index, column=21, value=sell.description)
            row_index += 1

        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['J'].width = 15         # Payment type
        sheet.column_dimensions['K'].width = 20         # Payment status
        sheet.column_dimensions['L'].width = 25         # Expense type
        sheet.column_dimensions['M'].width = 17         # Expense amount
        sheet.column_dimensions['N'].width = 15         # total
        sheet.column_dimensions['O'].width = 15         # gst
        sheet.column_dimensions['P'].width = 25         # total of other Expenses
        sheet.column_dimensions['Q'].width = 15         # Gross Total
        sheet.column_dimensions['R'].width = 15         # discount
        sheet.column_dimensions['S'].width = 15         # paid amount
        sheet.column_dimensions['T'].width = 15         # net Total
        sheet.column_dimensions['U'].width = 30         # Description

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Sell_invoice_data.xlsx"'
        wb.save(response)
        return response
    else:
        return redirect("/")

# To generate excel file of Purchase invoice data
def generate_excel_for_purchaseBills(request):
    if 'email' in request.session:
        PurchaseData = Purchases.objects.all()

        wb = Workbook()

        sheet = wb.active

        sheet['A1'] = 'BillNo.'                         #1
        sheet['B1'] = 'Bill Date.'                      #2
        sheet['C1'] = 'Supplier Name'                   #3
        sheet['D1'] = 'Supplier Phone'                  #4
        sheet['E1'] = 'Supplier GST'                    #5
        sheet['F1'] = 'Product Name'                    #6
        sheet['G1'] = 'Quantity'                        #7
        sheet['H1'] = 'Unit'                            #8
        sheet['I1'] = 'Price/Unit'                      #9
        sheet['J1'] = 'Payment Type'                    #10
        sheet['K1'] = 'Payment Status'                  #11
        sheet['L1'] = 'Expense Type'                    #12
        sheet['M1'] = 'Expense Amount'                  #13
        sheet['N1'] = 'Total'                           #14
        sheet['O1'] = 'GST (18%)'                       #15
        sheet['P1'] = 'Total Of Other Expenses'         #16
        sheet['Q1'] = 'Gross Total'                     #17
        sheet['R1'] = 'Discount'                        #18
        sheet['S1'] = 'Paid Amount'                     #19
        sheet['T1'] = 'Net Total'                       #20
        sheet['U1'] = 'Description'                     #21

        # Apply styles to the header row
        for cell in sheet[1]:
            cell.alignment = Alignment(horizontal='center')

        row_index = 2
        left = Alignment(horizontal='left')
        for index, purchase in enumerate(PurchaseData, start=1):
            formatted_date = purchase.billDate.strftime('%Y-%m-%d')
            sheet.cell(row=row_index, column=1, value=purchase.billNo).alignment = left
            sheet.cell(row=row_index, column=2, value=formatted_date)
            sheet.cell(row=row_index, column=3, value=purchase.supplierName)
            sheet.cell(row=row_index, column=4, value=purchase.supplierPhone)
            sheet.cell(row=row_index, column=5, value=purchase.supplierGst)

            product_column = 6

            purchased_products = PurchasedProducts.objects.filter(billNo=purchase.billNo)
            purchased_products_length = len(purchased_products)

            for i, product in enumerate(purchased_products, start=1):
                sheet.cell(row=row_index, column=product_column, value=product.product)
                sheet.cell(row=row_index, column=product_column + 1, value=product.quantity)
                sheet.cell(row=row_index, column=product_column + 2, value=product.unit)
                sheet.cell(row=row_index, column=product_column + 3, value=product.pricePerUnit)
                if i == purchased_products_length:
                    pass
                else:
                    row_index += 1
            sheet.cell(row=row_index, column=10, value=purchase.paymentType)
            sheet.cell(row=row_index, column=11, value=purchase.paymentStatus)

            other_expense = 12
            otherExpenses = OtherExpensesOfPurchase.objects.filter(billNo=purchase.billNo)
            otherExpenses_length = len(otherExpenses)
            for i, otherExpense in enumerate(otherExpenses, start=1):
                sheet.cell(row=row_index, column=other_expense, value=otherExpense.expenseType)
                sheet.cell(row=row_index, column=other_expense + 1, value=otherExpense.amount)
                if i == otherExpenses_length:
                    pass
                else:
                    row_index += 1
            sheet.cell(row=row_index, column=14, value=purchase.total)
            sheet.cell(row=row_index, column=15, value=purchase.gst)
            sheet.cell(row=row_index, column=16, value=purchase.otherExpenses)
            sheet.cell(row=row_index, column=17, value=purchase.grossTotal)
            sheet.cell(row=row_index, column=18, value=purchase.discount)
            sheet.cell(row=row_index, column=19, value=purchase.paidAmount)
            sheet.cell(row=row_index, column=20, value=purchase.netTotal)
            sheet.cell(row=row_index, column=21, value=purchase.description)
            row_index += 1

        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['J'].width = 15         # Payment type
        sheet.column_dimensions['K'].width = 20         # Payment status
        sheet.column_dimensions['L'].width = 25         # Expense type
        sheet.column_dimensions['M'].width = 17         # Expense amount
        sheet.column_dimensions['N'].width = 15         # total
        sheet.column_dimensions['O'].width = 15         # gst
        sheet.column_dimensions['P'].width = 25         # total of other Expenses
        sheet.column_dimensions['Q'].width = 15         # Gross Total
        sheet.column_dimensions['R'].width = 15         # discount
        sheet.column_dimensions['S'].width = 15         # paid amount
        sheet.column_dimensions['T'].width = 15         # net Total
        sheet.column_dimensions['U'].width = 30         # Description

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Purchase_invoice_data.xlsx"'
        wb.save(response)
        return response
    else:
        return redirect("/")

#------------- Support Section -----------------

def support(request):
    if 'email' in request.session:
        return render(request,"support.html")
    else:
        return redirect("/")